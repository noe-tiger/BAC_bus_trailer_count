import re
import argparse
import datetime
from openpyxl import load_workbook

WARNING_HUGE_DISTANCE = 2000
WARNING_LOW_DISTANCE = 10

def getData(ws, start, end):
  data = []
  for i in range(1, ws.max_row + 1):
    if ws[i][0].value != None:
      if ws[i][8].value != None and not 'ras' in ws[i][8].value.lower() and not 'aucun' in ws[i][8].value.lower():
        print('MESSAGE: (l %d) %s : %s' % (i, ws[i][1].value, ws[i][8].value))
      try:
        if ws[i][3].value > start['datetime'] and ws[i][3].value < end['datetime']:
          data.append({
            'index': i,
            'dateur': ws[i][0].value,
            'lieu': ws[i][1].value,
            'lieu_short': re.sub('[^A-Za-z -]+', '', ws[i][1].value).lower().split(' ')[0].split('-')[0], ## epur the string + get the first word
            'depart': ws[i][2].value,
            'date': ws[i][3].value,
            'vehicle': ws[i][4].value,
            'km': int(ws[i][5].value),
            'remorque': ws[i][6].value
          })
      except:
        pass
  return data

def correspond(start, end):
  if start['lieu_short'] == end['lieu_short'] and \
     start['date'] < end['date'] and \
     start['vehicle'] == end['vehicle'] and \
     start['remorque'] == end['remorque']:
    return True
  return False

def makePairs(data):
  pairs = []
  start = []
  end = []
  for elem in data:
    if elem['depart'] == "Départ en sortie":
      start.append(elem)
    else:
      end.append(elem)

  while len(start) and len(end):
    boul = True
    for i in range(len(start)):
      for j in range(len(end)):
        if correspond(start[i], end[j]):
          pairs.append({ "start": start[i], "end": end[j] })
          start.pop(i)
          end.pop(i)
          boul = False
        if not boul:
          break
      if not boul:
        break
    if boul:
      break
  
  print()
  print('-----------------')
  print()

  print("routes found : ", len(pairs))
  print("errors: ", len(start) + len(end))
  for elem in start:
    print("ERROR: no end: (line %d) %s" % (elem['index'], elem['lieu']))
  for elem in end:
    print("ERROR: no start: (line %d) %s" % (elem['index'], elem['lieu']))
  return pairs

def printDistance(pairs):

  print()
  print('-----------------')
  print()

  dist = {}
  camion = {}
  for elem in pairs:
    if elem['start']['vehicle'] not in dist:
      dist[elem['start']['vehicle']] = 0
    dist[elem['start']['vehicle']] += elem['end']['km'] - elem['start']['km']
    if elem['start']['remorque'] not in dist:
      dist[elem['start']['remorque']] = 0
    dist[elem['start']['remorque']] += elem['end']['km'] - elem['start']['km']
    if elem['end']['km'] - elem['start']['km'] > WARNING_HUGE_DISTANCE:
      print('WARNING: huge kilometers (l %d and l %d) %s : Got %s KM' % \
         (elem['start']['index'], elem['end']['index'], elem['start']['lieu'], elem['end']['km'] - elem['start']['km']))
    if elem['end']['km'] - elem['start']['km'] < WARNING_LOW_DISTANCE:
      print('WARNING: low kilometers (l %d and l %d) %s : Got %s KM' % \
         (elem['start']['index'], elem['end']['index'], elem['start']['lieu'], elem['end']['km'] - elem['start']['km']))

  print()
  print('-----------------')
  print()
  for elem in camion:
    print("%s did %d km" % (elem, camion[elem]))
  for elem in dist:
    print("%s did %d km" % (elem, dist[elem]))

def main(filename, start, end):
  wb = load_workbook(filename)
  ws = wb.worksheets[0]

  data = getData(ws, start, end)
  pairs = makePairs(data)
  printDistance(pairs)

def parse_date(date):
  split = date.split('/')
  if len(split) != 3:
    return None
  return {
    "datetime": datetime.datetime(int(split[2]), int(split[1]), int(split[0])),
    "full": date
  }

if __name__ == "__main__":
  start = None
  end = None

  parser = argparse.ArgumentParser(description='Get the trailer kms')
  parser.add_argument('-f', '--filename', type=str, metavar="", required=True, help='path to file to parse')
  parser.add_argument('-s', '--start', type=str, metavar="", help='date to start parsing (DD/MM/YYYY)')
  parser.add_argument('-e', '--end', type=str, metavar="", help='date to finish parsing (DD/MM/YYYY)')
  args = parser.parse_args()

  if args.start:
    start = parse_date(args.start)
    if start == None:
      parser.print_help()
      exit(-1)
  else:
    start = { "datetime": datetime.datetime.min }
  if args.end:
    end = parse_date(args.end)
    if end == None:
      parser.print_help()
      exit(-1)
  else:
    end = { "datetime": datetime.datetime.max }
  main(args.filename, start, end)